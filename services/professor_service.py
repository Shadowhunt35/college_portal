"""
Professor Service
Handles all professor-related business logic.
"""

import datetime
from extensions import db
from models import (User, Subject, Attendance, Mark,
                    CGPA, ProfessorSubject)
from ml.predict import predict_student_risk
from src.logger import logger

import io
import pandas as pd



def get_professor_subjects(professor: User) -> list:
    """Get all subjects assigned to this professor."""
    assignments = ProfessorSubject.query.filter_by(
        professor_id=professor.id
    ).all()
    return assignments


def get_subject_students(subject_id: int, batch_id: int) -> list:
    """Get all students for a given subject and batch."""
    assignment = ProfessorSubject.query.filter_by(
        subject_id=subject_id,
        batch_id=batch_id
    ).first()

    if not assignment:
        return []

    subject = Subject.query.get(subject_id)
    if not subject:
        return []

    students = User.query.filter_by(
        role          = 'student',
        department_id = subject.department_id,
        batch_id      = batch_id,
        current_semester = subject.semester
    ).order_by(User.reg_no).all()

    return students


def get_professor_dashboard_data(professor: User) -> dict:
    """Get all data for professor dashboard."""
    assignments = get_professor_subjects(professor)

    subject_stats = []
    for a in assignments:
        subject  = a.subject
        students = get_subject_students(subject.id, a.batch_id)

        # Count attendance records for today
        today = datetime.datetime.today().date()
        marked_today = Attendance.query.filter_by(
            subject_id = subject.id,
            date       = today
        ).count()

        # Count marks entered
        marks_entered = Mark.query.filter_by(
            subject_id = subject.id
        ).count()

        subject_stats.append({
            'assignment':    a,
            'subject':       subject,
            'batch':         a.batch,
            'student_count': len(students),
            'marked_today':  marked_today > 0,
            'marks_entered': marks_entered,
        })

    return {
        'assignments':   assignments,
        'subject_stats': subject_stats,
    }


def get_attendance_data(professor: User, subject_id: int, batch_id: int) -> dict:
    """Get attendance records for a subject."""
    # Verify professor owns this subject
    assignment = ProfessorSubject.query.filter_by(
        professor_id = professor.id,
        subject_id   = subject_id,
        batch_id     = batch_id
    ).first()

    if not assignment:
        return {'error': 'You are not assigned to this subject.'}

    subject  = Subject.query.get(subject_id)
    students = get_subject_students(subject_id, batch_id)

    # Get all attendance dates for this subject
    all_records = Attendance.query.filter_by(
        subject_id=subject_id
    ).order_by(Attendance.date).all()

    # Unique dates
    dates = sorted(set(r.date for r in all_records))

    # Build attendance matrix: {student_id: {date: present}}
    matrix = {}
    for student in students:
        matrix[student.id] = {}
        for date in dates:
            record = Attendance.query.filter_by(
                student_id = student.id,
                subject_id = subject_id,
                date       = date
            ).first()
            matrix[student.id][date] = record.present if record else None

    return {
        'subject':  subject,
        'students': students,
        'dates':    dates,
        'matrix':   matrix,
        'batch':    assignment.batch,
    }


def mark_attendance(professor: User, subject_id: int,
                    batch_id: int, date_str: str,
                    att_type: str, present_ids: list,
                    academic_year: str) -> dict:
    """
    Save attendance for a list of students.
    present_ids = list of student IDs marked present.
    All other students in the class = absent.
    """
    # Verify professor owns this subject
    assignment = ProfessorSubject.query.filter_by(
        professor_id = professor.id,
        subject_id   = subject_id,
        batch_id     = batch_id
    ).first()

    if not assignment:
        return {'success': False, 'error': 'Not authorised for this subject.'}

    try:
        date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        return {'success': False, 'error': 'Invalid date format.'}

    students = get_subject_students(subject_id, batch_id)

    for student in students:
        present = student.id in present_ids
        record  = Attendance.query.filter_by(
            student_id = student.id,
            subject_id = subject_id,
            date       = date,
            type       = att_type
        ).first()

        if record:
            record.present = present
        else:
            db.session.add(Attendance(
                student_id    = student.id,
                subject_id    = subject_id,
                date          = date,
                type          = att_type,
                present       = present,
                academic_year = academic_year
            ))

    db.session.commit()
    logger.info(
        f'Attendance marked by {professor.reg_no} for '
        f'subject {subject_id} on {date_str} ({att_type})')

    return {
        'success': True,
        'message': f'Attendance saved for {date_str} ({att_type}).'
    }


def get_marks_data(professor: User, subject_id: int, batch_id: int) -> dict:
    """Get marks for all students in a subject."""
    assignment = ProfessorSubject.query.filter_by(
        professor_id = professor.id,
        subject_id   = subject_id,
        batch_id     = batch_id
    ).first()

    if not assignment:
        return {'error': 'You are not assigned to this subject.'}

    subject  = Subject.query.get(subject_id)
    students = get_subject_students(subject_id, batch_id)

    student_marks = []
    for student in students:
        mark = Mark.query.filter_by(
            student_id    = student.id,
            subject_id    = subject_id,
            academic_year = assignment.academic_year
        ).first()

        student_marks.append({
            'student': student,
            'mark':    mark,
        })

    return {
        'subject':       subject,
        'student_marks': student_marks,
        'batch':         assignment.batch,
        'academic_year': assignment.academic_year,
    }


def save_marks(professor: User, subject_id: int,
               batch_id: int, academic_year: str,
               marks_data: list) -> dict:
    """
    Save marks for multiple students.
    marks_data = list of dicts with student_id + mark fields.
    """
    assignment = ProfessorSubject.query.filter_by(
        professor_id = professor.id,
        subject_id   = subject_id,
        batch_id     = batch_id
    ).first()

    if not assignment:
        return {'success': False, 'error': 'Not authorised for this subject.'}

    subject = Subject.query.get(subject_id)
    saved   = 0
    errors  = []

    for entry in marks_data:
        student_id = entry.get('student_id')
        try:
            theory   = _clamp(entry.get('theory_internal'), 0, 20)
            asgn     = _clamp(entry.get('assignment'), 0, 5)
            att_mark = _clamp(entry.get('attendance_marks'), 0, 5)
            lab      = _clamp(entry.get('lab_internal'), 0, 20) if subject.has_lab else None
        except (TypeError, ValueError):
            errors.append(f'Invalid marks for student ID {student_id}')
            continue

        mark = Mark.query.filter_by(
            student_id    = student_id,
            subject_id    = subject_id,
            academic_year = academic_year
        ).first()

        if mark:
            mark.theory_internal  = theory
            mark.assignment       = asgn
            mark.attendance_marks = att_mark
            mark.lab_internal     = lab
        else:
            db.session.add(Mark(
                student_id       = student_id,
                subject_id       = subject_id,
                academic_year    = academic_year,
                theory_internal  = theory,
                assignment       = asgn,
                attendance_marks = att_mark,
                lab_internal     = lab
            ))
        saved += 1

    db.session.commit()
    logger.info(
        f'Marks saved by {professor.reg_no} for '
        f'subject {subject_id} — {saved} students')

    return {
        'success': True,
        'saved':   saved,
        'errors':  errors,
        'message': f'Marks saved for {saved} students.' +
                   (f' {len(errors)} errors.' if errors else '')
    }


def get_risk_report(professor: User) -> list:
    """Risk report for all students across professor's subjects."""
    assignments = get_professor_subjects(professor)
    report      = []
    seen        = set()

    for a in assignments:
        students = get_subject_students(a.subject_id, a.batch_id)

        for student in students:
            if student.id in seen:
                continue
            seen.add(student.id)

            # Compute averages
            marks   = Mark.query.filter_by(student_id=student.id).all()
            att_all = Attendance.query.filter_by(student_id=student.id).all()
            cgpa    = CGPA.query.filter_by(student_id=student.id)\
                              .order_by(CGPA.semester.desc()).first()

            att_pct = (
                round(sum(1 for r in att_all if r.present) / len(att_all) * 100, 1)
                if att_all else 0
            )
            avg_theory = (
                round(sum(m.theory_internal or 0 for m in marks) / len(marks), 1)
                if marks else 0
            )
            avg_asgn = (
                round(sum(m.assignment or 0 for m in marks) / len(marks), 1)
                if marks else 0
            )
            avg_att_marks = (
                round(sum(m.attendance_marks or 0 for m in marks) / len(marks), 1)
                if marks else 0
            )

            risk = predict_student_risk(
                attendance_pct   = att_pct,
                theory_internal  = avg_theory,
                assignment       = avg_asgn,
                attendance_marks = avg_att_marks,
                cgpa             = cgpa.value if cgpa else 0
            )

            report.append({
                'student':    student,
                'att_pct':    att_pct,
                'avg_theory': avg_theory,
                'cgpa':       cgpa.value if cgpa else 'N/A',
                'risk':       risk,
            })

    report.sort(key=lambda x: x['risk']['score'])
    return report


def save_cgpa(professor: User, student_id: int,
              semester: int, value: float,
              academic_year: str) -> dict:
    """Save or update a student's CGPA for a semester."""
    try:
        value = float(value)
        if not (0 <= value <= 10):
            return {'success': False, 'error': 'CGPA must be between 0 and 10.'}
    except (TypeError, ValueError):
        return {'success': False, 'error': 'Invalid CGPA value.'}

    record = CGPA.query.filter_by(
        student_id    = student_id,
        semester      = semester,
        academic_year = academic_year
    ).first()

    if record:
        record.value = value
    else:
        db.session.add(CGPA(
            student_id    = student_id,
            semester      = semester,
            value         = value,
            academic_year = academic_year
        ))

    db.session.commit()
    logger.info(f'CGPA saved for student {student_id} sem {semester}: {value}')
    return {'success': True}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clamp(value, min_val, max_val):
    """Clamp a numeric value between min and max."""
    if value is None or value == '':
        return None
    return max(min_val, min(max_val, float(value)))

"""
download attendance report as Excel for a date range
"""




"""
REPLACE the get_attendance_report function in services/professor_service.py
"""

def get_attendance_report(professor, subject_id: int, batch_id: int,
                          from_date: str, to_date: str, att_type: str = 'both'):
    """
    Generate Excel attendance report with full header info.
    Returns (bytes, subject, start, end) or None on error.
    """
    import io
    import datetime
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    assignment = ProfessorSubject.query.filter_by(
        professor_id=professor.id,
        subject_id=subject_id,
        batch_id=batch_id
    ).first()
    if not assignment:
        return None

    subject  = Subject.query.get(subject_id)
    students = get_subject_students(subject_id, batch_id)
    batch    = assignment.batch

    try:
        start = datetime.date.fromisoformat(from_date)
        end   = datetime.date.fromisoformat(to_date)
    except (ValueError, TypeError):
        return None

    # Fetch records
    query = Attendance.query.filter(
        Attendance.subject_id == subject_id,
        Attendance.date >= start,
        Attendance.date <= end,
    )
    if att_type != 'both':
        query = query.filter(Attendance.type == att_type)

    records  = query.order_by(Attendance.date, Attendance.type).all()
    dates    = sorted(set((r.date, r.type) for r in records))
    rec_map  = {(r.student_id, r.date, r.type): r.present for r in records}

    # Build data rows
    rows = []
    for student in students:
        row = {'Roll No': student.reg_no, 'Name': student.name}
        present_count = total_count = 0
        for (date, rtype) in dates:
            col = f'{date.strftime("%d-%b")} ({rtype[0].upper()})'
            val = rec_map.get((student.id, date, rtype))
            if val is True:
                row[col] = 'P'; present_count += 1; total_count += 1
            elif val is False:
                row[col] = 'A'; total_count += 1
            else:
                row[col] = '-'
        row['Present'] = present_count
        row['Total']   = total_count
        row['Att. %']  = round(present_count / total_count * 100, 1) if total_count else 0
        rows.append(row)

    df = pd.DataFrame(rows)

    # Write to Excel with rich header
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write data starting from row 7 (leave 6 rows for header)
        df.to_excel(writer, index=False, sheet_name='Attendance', startrow=6)
        ws = writer.sheets['Attendance']

        # ── Info header block ──────────────────────────────────────
        header_fill  = PatternFill("solid", fgColor="4A2518")   # MCE brown
        subhead_fill = PatternFill("solid", fgColor="6B3A2A")
        white_font   = Font(color="FFFFFF", bold=True, size=12)
        normal_font  = Font(color="FFFFFF", size=10)
        center       = Alignment(horizontal="center", vertical="center")

        # Row 1: College name
        ws.merge_cells('A1:H1')
        ws['A1'] = 'MOTIHARI COLLEGE OF ENGINEERING, MOTIHARI'
        ws['A1'].font      = Font(color="FFFFFF", bold=True, size=14)
        ws['A1'].fill      = header_fill
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 28

        # Row 2: Affiliation
        ws.merge_cells('A2:H2')
        ws['A2'] = 'Bihar Engineering University | AICTE Approved'
        ws['A2'].font      = Font(color="F5C842", size=10)
        ws['A2'].fill      = header_fill
        ws['A2'].alignment = center
        ws.row_dimensions[2].height = 18

        # Row 3: blank separator
        ws.merge_cells('A3:H3')
        ws['A3'].fill = subhead_fill
        ws.row_dimensions[3].height = 6

        # Row 4: Subject + Dept
        ws.merge_cells('A4:D4')
        ws['A4'] = f'Subject: {subject.name} ({subject.code})'
        ws['A4'].font = normal_font
        ws['A4'].fill = subhead_fill
        ws['A4'].alignment = Alignment(vertical="center", indent=1)

        ws.merge_cells('E4:H4')
        dept_name = subject.department.name if subject.department else 'N/A'
        ws['E4'] = f'Department: {dept_name}'
        ws['E4'].font = normal_font
        ws['E4'].fill = subhead_fill
        ws['E4'].alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[4].height = 20

        # Row 5: Batch + Professor
        ws.merge_cells('A5:D5')
        ws['A5'] = f'Batch: {batch.name} ({batch.start_year}–{batch.end_year})'
        ws['A5'].font = normal_font
        ws['A5'].fill = subhead_fill
        ws['A5'].alignment = Alignment(vertical="center", indent=1)

        ws.merge_cells('E5:H5')
        ws['E5'] = f'Professor: {professor.name}'
        ws['E5'].font = normal_font
        ws['E5'].fill = subhead_fill
        ws['E5'].alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[5].height = 20

        # Row 6: Date range + Type
        ws.merge_cells('A6:D6')
        ws['A6'] = f'Period: {start.strftime("%d %b %Y")} to {end.strftime("%d %b %Y")}'
        ws['A6'].font = normal_font
        ws['A6'].fill = subhead_fill
        ws['A6'].alignment = Alignment(vertical="center", indent=1)

        ws.merge_cells('E6:H6')
        ws['E6'] = f'Type: {att_type.capitalize()} | Semester: {subject.semester}'
        ws['E6'].font = normal_font
        ws['E6'].fill = subhead_fill
        ws['E6'].alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[6].height = 20

        # ── Style the data header row (row 7) ─────────────────────
        col_header_fill = PatternFill("solid", fgColor="C9983A")
        for cell in ws[7]:
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = col_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[7].height = 22

        # ── Color P/A cells ────────────────────────────────────────
        green_fill = PatternFill("solid", fgColor="C8F7C5")
        red_fill   = PatternFill("solid", fgColor="FADBD8")
        for row in ws.iter_rows(min_row=8):
            for cell in row:
                if cell.value == 'P':
                    cell.fill      = green_fill
                    cell.alignment = Alignment(horizontal="center")
                elif cell.value == 'A':
                    cell.fill      = red_fill
                    cell.alignment = Alignment(horizontal="center")
                elif cell.column == 1:
                    cell.font = Font(bold=True)

        # ── Auto-size columns ──────────────────────────────────────
        

        for i, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass

        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 20)

    output.seek(0)
    return output.getvalue(), subject, start, end